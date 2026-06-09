from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

LAST_UPDATED = "2026-06-08"

COLUMNS = ["#", "Area", "Item", "Description", "Priority", "Status", "Notes"]
STATUSES = ["Done", "In Progress", "Not Started", "Blocked"]
PRIORITIES = ["P0", "P1", "P2", "P3"]

ITEMS = [
    [1, "Screens", "Dashboard", "Live opacity and runtime taskbar control screen.", "P0", "Done", "Refined from generated dashboard reference."],
    [2, "Screens", "Tuning", "Material and user settings tuning screen.", "P0", "Done", "Clear, Glass, Solid presets plus transition, sensor, startup, tray, and hotkey settings."],
    [3, "Screens", "Monitors", "Per-display override management screen.", "P1", "Done", "Native taskbar-backed monitor catalog added."],
    [4, "Screens", "Automation", "State rules for hover, focus, maximized, and fullscreen.", "P0", "Done", "Runtime sensors now drive desktop, visible, maximized, fullscreen, and hover states."],
    [5, "Screens", "Diagnostics", "Runtime apply and state simulation screen.", "P0", "Done", "Manual runtime checks available."],
    [6, "Features", "Taskbar interop", "Apply Windows composition attributes to taskbar windows.", "P0", "Done", "Primary and secondary taskbar handles supported."],
    [7, "Features", "Persistence", "Save profile and behavior settings locally.", "P0", "Done", "JSON settings store implemented."],
    [8, "Features", "Tray integration", "Background tray menu and open/apply/toggle/tuning/exit commands.", "P1", "Done", "Implemented with Shell_NotifyIcon Win32 host to keep WinUI project shape."],
    [9, "Features", "Global hotkeys", "Register open and toggle shortcuts.", "P1", "Done", "Ctrl+Alt+G and Ctrl+Alt+T register against the WinUI window handle."],
    [10, "Backend", "Opacity policy", "Resolve opacity from profile and runtime state.", "P0", "Done", "Covered by unit tests."],
    [11, "Backend", "Focus/fullscreen sensors", "Detect active window, maximized, fullscreen, and hover taskbar states.", "P1", "Done", "Win32 foreground, monitor, taskbar, and cursor sensors apply policy on state changes."],
    [12, "Infrastructure", "Tests", "Focused unit tests for core policy behavior.", "P0", "Done", "Profile migration, tuning, appearance, sensors, hotkeys, policy, and snapshot tests covered."],
    [13, "Infrastructure", "Launcher", "Build and launch script with log cleanup.", "P0", "Done", "Batch launcher added."],
    [14, "Infrastructure", "Release snapshot", "SemVer snapshot branch/tag publishing script.", "P0", "Done", "Script added for new remote."],
    [15, "Infrastructure", "Remote repository", "Own GitHub remote for WinUI rebuild.", "P0", "Done", "Remote created and main/snapshots pushed."],
    [16, "Screens", "Generated mockups", "Single-screen generated images for key views.", "P1", "Done", "V1 P1 mockup batch organized into logical screen folders."],
    [17, "Infrastructure", "Release build hygiene", "Keep Release x64 build free of app warnings.", "P0", "Done", "Source-generated settings JSON and trimming disabled."],
    [18, "Screens", "Dashboard visual polish", "Match Dashboard layout to generated Oxygen Taskbar direction.", "P0", "Done", "Dark shell, KPI cards, opacity rail, segmented actions, and runtime rows."],
    [19, "Infrastructure", "Self-contained launch", "Launch unpackaged WinUI without requiring registered Windows App Runtime.", "P0", "Done", "Windows App SDK self-contained settings added and launcher verified."],
    [20, "Screens", "V1 P1 mockup prompt handoff", "Prompt next batch of supporting app states and edge cases.", "P1", "Done", "Covers onboarding, customization, monitor detail, automation builder, diagnostics error, hotkeys, tray, and about/update."],
    [21, "Screens", "Onboarding first-run mockup", "Generated first-run setup image organized under SCREEN IMAGES/Onboarding.", "P1", "Done", "Validated as a single standalone full-screen mockup."],
    [22, "Screens", "V1 P1 mockup organization", "Organize generated supporting app states and edge cases.", "P1", "Done", "Presets, monitors, automation, diagnostics, settings, tray, about, and alternate onboarding states organized."],
    [23, "Screens", "V1 P1 screen implementation", "Implement supporting app states from the organized V1 P1 mockups.", "P1", "Done", "First-run, preset editor, monitor detail, rule builder, diagnostics error, hotkeys, and about/update surfaces added."],
    [24, "Features", "First-run setup flow", "Route new installs through a local-only onboarding screen before the dashboard.", "P1", "Done", "Start with Oxygen Clear persists completion and enters the dashboard."],
    [25, "Features", "Startup registration", "Persist Start with Windows through the current-user Run key.", "P1", "Done", "Settings toggle writes and removes the OxygenTaskbar startup value."],
    [26, "Backend", "Monitor catalog expansion", "Discover taskbar-backed monitor profiles from primary and secondary taskbar windows.", "P1", "Done", "Maps Shell_TrayWnd and Shell_SecondaryTrayWnd windows to monitor device names."],
    [27, "Backend", "Runtime automation sensors", "Continuously classify runtime desktop/window/fullscreen/hover state.", "P1", "Done", "Timer-backed sensor applies policy only when the resolved trigger changes."],
    [28, "Screens", "V1 P2 mockup prompt handoff", "Prepare next generated-image part for post-runtime behavior screens.", "P1", "Done", "Active reference part created for history, calibration, conflicts, recovery, and multi-monitor telemetry states."],
    [29, "Screens", "V1 P2 mockup organization", "Organize generated post-runtime behavior screens.", "P1", "Done", "History, calibration, conflict, diagnostics, monitor, settings, and dashboard mockups organized."],
    [30, "Screens", "Runtime dashboard implementation", "Implement Dashboard runtime sensor active state from V1 P2.", "P1", "Done", "Dashboard shows active sensor, resolved opacity, automation status, and recent runtime history."],
    [31, "Screens", "Automation history and calibration", "Implement live history, sensor calibration, and conflict warning states.", "P1", "Done", "Automation page shows runtime history, hover calibration, sensor status, and paused-rule recovery."],
    [32, "Screens", "Diagnostics recovery states", "Implement sensor timeline and hotkey conflict recovery views.", "P1", "Done", "Diagnostics shows sensor timeline, hotkey status, and reset recovery alongside runtime details."],
    [33, "Screens", "Multi-display overview", "Implement monitor overview with multi-display state and recent actions.", "P1", "Done", "Monitor page shows display KPIs, detected display rows, sync state, and recent monitor actions."],
    [34, "Screens", "Startup permission warning", "Implement startup registration failure recovery state.", "P1", "Done", "Tuning shows startup registration status, permission warning copy, and retry recovery."],
    [35, "Features", "Whole taskbar transparency parity", "Match old repo behavior by fading taskbar windows and icons, not only accent material.", "P0", "Done", "Taskbar interop now applies WS_EX_LAYERED and SetLayeredWindowAttributes after material composition."],
    [36, "Features", "Fade transition controls", "Expose fade duration and easing settings and animate taskbar opacity changes.", "P1", "Done", "Tuning now saves fade speed/easing and taskbar alpha changes animate with cancellable easing."],
    [37, "Features", "Unified tuning settings", "Move user-set controls into renamed Tuning tab and keep runtime history in Diagnostics.", "P1", "Done", "Tuning now owns fade-in/out, hover proximity, automation sensors, tray/startup, and hotkey settings; Diagnostics keeps telemetry/history."],
    [38, "Infrastructure", "Transition settings migration", "Preserve existing fade timing when loading settings created before fade-in/out split.", "P1", "Done", "Legacy profiles with fadeMilliseconds now hydrate fade-in and fade-out instead of becoming instant."],
    [39, "Features", "Tuning save edge cases", "Ensure Save changes preserves the active material and visible profile values.", "P1", "Done", "Saving from Oxygen Clear no longer silently rebuilds the profile from Focus Glass; blank names preserve the current profile name."],
    [40, "Backend", "Hover proximity calibration", "Ensure the Tuning hover proximity control drives runtime hover detection.", "P1", "Done", "Runtime sensor now uses the saved hover distance instead of a hard-coded 10 px threshold, with boundary coverage."],
    [41, "Features", "Close-to-tray lifecycle", "Keep the app running when the dashboard is closed while the tray icon is enabled.", "P1", "Done", "Close now hides the WinUI window, tray and hotkey requests restore it, and tray Open Tuning routes to the Tuning page."],
    [42, "Screens", "Ultrawide layout readability", "Keep primary app pages readable on widescreen monitors.", "P1", "Done", "Dashboard, Tuning, Monitors, Automation, Diagnostics, About, and Settings page content is centered with a readable max width."],
    [43, "Features", "Hotkey registration diagnostics", "Report actual Windows hotkey registration success instead of only validating shortcut text.", "P1", "Done", "Diagnostics now distinguishes registered shortcuts, invalid formats, and Windows registration failures with error codes."],
    [44, "Features", "Tray command dispatch audit", "Verify tray menu labels and command routing for dashboard, tuning, apply, toggle, and exit.", "P1", "Done", "Tray command labels and dispatch are covered by focused tests; live hidden-window validation confirmed background toggle and dashboard restore outcomes."],
    [45, "Infrastructure", "Settings persistence hardening", "Protect local settings writes and verify relaunch persistence.", "P1", "Done", "SettingsStore now writes atomically; tests cover round-trip, corrupt fallback, and legacy fade migration; Computer Use confirmed Tuning sensor settings persist on a 3440-wide relaunch."],
    [46, "Features", "Tuning profile name edit stability", "Preserve unsaved profile name edits while other Tuning settings refresh.", "P1", "Done", "Tuning no longer overwrites dirty preset-name input during state refresh; Computer Use reproduced the bug and verified save plus relaunch on a 3440-wide monitor."],
    [47, "Features", "Tuning material and reset edge cases", "Ensure material buttons and Reset behave intentionally from the Tuning screen.", "P1", "Done", "Clear, Acrylic, Mica, and Solid are wired; material changes preserve current tuning values; Reset restores Oxygen Clear; Computer Use verified Solid persistence and Reset on a 3440-wide monitor."],
    [48, "Features", "Tuning slider refresh stability", "Preserve unsaved opacity, fade, and easing edits while other Tuning settings refresh.", "P1", "Done", "Computer Use reproduced hover proximity refresh wiping dirty opacity/fade edits, then verified the dirty tuning values survive on a 3440-wide monitor."],
    [49, "Features", "Transparency toggle baseline stability", "Reset stale transparency-toggle restore state when the user applies a new profile or explicit opacity.", "P1", "Done", "Computer Use reproduced Night Solid inheriting an old 32% toggle baseline after a transparent state, then verified profile changes establish a fresh toggle baseline."],
    [50, "Features", "Automation policy preview accuracy", "Show actual automation rule targets and live runtime preview instead of static mock values.", "P1", "Done", "Computer Use reproduced the 72/85/60/42/Hidden mismatch and verified policy-driven values on a 3440-wide monitor."],
    [51, "Features", "Automation paused recovery", "Let the Automation conflict card re-enable paused automation directly.", "P1", "Done", "Computer Use reproduced the disabled paused-state recovery button and verified Enable automation now restores live automation on a 3440-wide monitor."],
    [52, "Features", "Diagnostics recovery action accuracy", "Keep Diagnostics recovery controls truthful and non-destructive.", "P1", "Done", "Computer Use reproduced the inert Open folder button, misleading error header, and fullscreen Apply defaults behavior; Diagnostics now opens logs, reports healthy taskbar updates, and applies safe desktop without changing saved settings."],
    [53, "Features", "Tuning hotkey editing", "Make the unified Tuning hotkey fields actually editable and persist them through Windows hotkey registration.", "P1", "Done", "Computer Use reproduced read-only hotkey fields on the Tuning page; Save changes now writes and re-registers edited shortcuts without overwriting opacity or transition values, with initialization guards added to live-control pages."],
    [54, "Features", "Settings hotkey warning accuracy", "Keep the legacy Settings hotkey status truthful if the surface is reached.", "P2", "Done", "Replaced the always-visible shortcut conflict warning with the same real Windows registration state used by Diagnostics."],
    [55, "Infrastructure", "Mockup handoff documentation freshness", "Remove stale active-prompt claims after generated mockups are organized.", "P2", "Done", "Computer Use swept the live app surfaces, then README and walkthrough were updated so V1 P2 prompt handoff is no longer described as active."],
    [56, "Features", "Automation preview affordance clarity", "Make Automation read like an intentional policy preview instead of a disabled editor.", "P2", "Done", "Computer Use reproduced disabled Add custom rule, reset, discard, and save controls; Automation now uses read-only policy output with an active Open Tuning edit path."],
    [57, "Features", "Diagnostics recovery card state", "Keep the Diagnostics recovery card aligned with healthy versus failed taskbar detection.", "P1", "Done", "Computer Use found a healthy Diagnostics page still showing probable failure causes; recovery copy is now state-aware and tested."],
    [58, "Features", "About page version and utility actions", "Keep About version labels and utility buttons truthful for local snapshot builds.", "P1", "Done", "Computer Use found hard-coded 1.0.0 version labels and inert utility buttons; About now reads assembly metadata, opens logs/settings, and reports snapshot update status."],
    [59, "Features", "Monitor override apply persistence", "Make monitor override controls apply to the selected display instead of the global active profile.", "P1", "Done", "Computer Use found monitor Apply reverting edited override values and recording a global desktop apply; monitor overrides now persist per display and feed taskbar opacity resolution."],
    [60, "Features", "Per-monitor transition direction audit", "Ensure fade-in and fade-out speeds are honored independently when different taskbars change in different directions.", "P1", "Done", "Code audit found multi-monitor animation used the first taskbar as the representative duration; each taskbar now selects its own fade direction, while unchanged taskbars skip no-op animation time."],
    [61, "Features", "Hover proximity boundary audit", "Verify hover proximity edge settings behave intentionally at 0 px and 48 px.", "P1", "Done", "Computer Use verified Tuning accepts and persists 0 px and 48 px hover proximity; sensor tests now lock zero and negative-distance boundary behavior."],
    [62, "Features", "Runtime state vocabulary consistency", "Keep runtime state labels consistent across Dashboard, Automation, Monitors, and Diagnostics.", "P1", "Done", "Source audit found Dashboard, Automation, Monitors, and Diagnostics using different names for the same runtime states; pages now share one runtime vocabulary with tests."],
    [63, "Features", "Manual apply semantics clarity", "Make manual runtime actions describe whether they reapply the current state or simulate a specific state.", "P1", "Done", "Audit found Dashboard and tray Apply forced Desktop while Automation preview simulated Visible window; controls now say and do Reapply current state or Simulate visible window explicitly."],
    [64, "Features", "Close-to-tray lifecycle audit", "Verify closing the app with the tray icon enabled hides the window without exiting the background service.", "P1", "Done", "Computer Use closed the Oxygen window; process inspection confirmed TaskbarTransparency stayed alive with no visible main window and ShowTrayIcon enabled."],
    [65, "Features", "Hidden-window hotkey restore audit", "Verify the global open hotkey restores Oxygen after the dashboard has been hidden to tray.", "P1", "Done", "Computer Use closed Oxygen to tray on a 1426 by 913 widescreen app surface, sent Ctrl+Alt+G through Windows, and confirmed the Dashboard window returned at the same bounds."],
    [66, "Features", "Tray popup command audit", "Verify the notification-area popup exposes the intended commands and routes key commands from the hidden tray state.", "P1", "Done", "Windows tray overflow exposed Oxygen Taskbar; native popup extraction confirmed Open Dashboard, Reapply Current State, Toggle Transparency, Open Tuning, and Exit labels. Open Tuning restored Tuning, and Exit terminated the background process."],
    [67, "Infrastructure", "Runtime smoothness optimization", "Reduce redundant disk writes and overlapping runtime sensor work during frequent taskbar applies.", "P1", "Done", "Runtime applies now update transient state without saving unchanged settings; identical settings saves are skipped; sensor ticks avoid overlap."],
    [68, "Infrastructure", "Taskbar interop no-op optimization", "Avoid redundant native taskbar composition and alpha calls when visual requests are unchanged.", "P1", "Done", "Taskbar appearance requests are cached per handle and unchanged layered-alpha requests are skipped, reducing work during rapid controls and repeated applies."],
    [69, "Infrastructure", "Slider persistence debounce", "Collapse rapid opacity and hover-distance slider bursts into immediate previews plus one settled settings save.", "P1", "Done", "Dashboard opacity and Tuning hover proximity now preview instantly while debouncing durable settings persistence to reduce drag-time disk churn."],
    [70, "Infrastructure", "Active page refresh optimization", "Limit global state refresh work to loaded pages and avoid replacing unchanged runtime list sources.", "P1", "Done", "Dashboard, Tuning, Automation, Monitors, Diagnostics, and Settings now unsubscribe when unloaded; runtime-heavy lists update only when their source signatures change."],
    [71, "Infrastructure", "Startup monitor refresh no-op guard", "Avoid startup settings writes and live monitor collection resets when detected monitors match saved state.", "P1", "Done", "Monitor refresh now merges detected display metadata with saved overrides, updates live monitors only when changed, and saves only when the effective monitor list changes."],
    [72, "Infrastructure", "Shared taskbar window discovery", "Centralize native taskbar enumeration and monitor metadata lookup used by sensors, monitor catalog, and appearance apply.", "P1", "Done", "TaskbarWindowCatalog now owns Shell_TrayWnd and Shell_SecondaryTrayWnd discovery so services reuse one native lookup path instead of duplicating Win32 enumeration code."],
    [73, "Infrastructure", "Taskbar animation cache cleanup", "Clean up stale animation/taskbar caches and completed animation cancellation tokens.", "P1", "Done", "Appearance apply now prunes handles no longer present in the live taskbar set and clears completed animation cancellation state without waiting for the next animation."],
    [74, "Infrastructure", "AppState no-op setter guards", "Avoid redundant save, notify, apply, tray, startup, and hotkey work when requested settings already match current state.", "P1", "Done", "State setters now short-circuit duplicate values while preserving debounced preview commits and retry paths for failed startup or hotkey registration."],
    [75, "Infrastructure", "Same-page navigation no-op guard", "Avoid recreating the current WinUI page when tray, hotkey, or navigation requests target the page already displayed.", "P1", "Done", "Main window navigation now skips same-page requests so loaded pages keep their existing controls, timers, and subscriptions instead of churning through unload/reload work."],
    [76, "Infrastructure", "Runtime event refresh versioning", "Avoid rebuilding runtime history signature strings on every Dashboard, Monitors, and Diagnostics refresh.", "P1", "Done", "RuntimeSnapshot now exposes a RecentEventsVersion counter so pages use integer change checks before rebuilding history list item sources."],
    [77, "Infrastructure", "Monitor list refresh versioning", "Avoid rebuilding monitor overview signature strings on every Monitors page refresh.", "P1", "Done", "AppState now tracks MonitorsVersion so monitor overview rows rebuild only when the live monitor collection or overrides change."],
    [78, "Infrastructure", "Loaded page refresh coalescing", "Avoid queuing duplicate UI refreshes for loaded pages during quick AppState change bursts.", "P1", "Done", "State-listening pages now share a RefreshCoalescer so repeated change notifications collapse to one pending dispatcher refresh per page."],
    [79, "Infrastructure", "Synced monitor apply lookup skip", "Avoid building per-monitor override lookup dictionaries when all monitors follow the primary opacity.", "P1", "Done", "TaskbarAppearanceService now creates the monitor override lookup only when at least one monitor is unsynced."],
    [80, "Infrastructure", "Taskbar apply diagnostics counters", "Expose lightweight counters for native composition skips, alpha skips, monitor lookup use, and animation starts.", "P1", "Done", "Diagnostics now shows latest apply counters so future optimization passes can be guided by measured skip behavior instead of source inspection alone."],
    [81, "Infrastructure", "Layered alpha change loop allocation reduction", "Avoid LINQ iterator and dictionary projection overhead while collecting changed layered-alpha targets.", "P1", "Done", "TaskbarAppearanceService now builds alpha-change targets in one pre-sized loop while preserving skip counters for Diagnostics."],
    [82, "Infrastructure", "Taskbar target dedupe allocation reduction", "Avoid LINQ grouping when de-duplicating taskbar targets by window handle before apply.", "P1", "Done", "TaskbarAppearanceService now preserves first target per handle with a single loop and pre-sizes per-apply target dictionaries."],
    [83, "Infrastructure", "Stale taskbar cache pruning allocation reduction", "Avoid LINQ projection and stale-handle arrays while pruning cached taskbar state after apply.", "P1", "Done", "TaskbarAppearanceService now prunes alpha and appearance caches directly against one live handle set built from current taskbar targets."],
    [84, "Infrastructure", "Monitor override lookup loop allocation reduction", "Avoid double-scanning and LINQ grouping when building per-monitor override lookup data.", "P1", "Done", "TaskbarAppearanceService now builds unsynced monitor override lookup entries in one loop and returns no dictionary for fully synced monitor sets."],
    [85, "Infrastructure", "Taskbar catalog iterator allocation reduction", "Avoid iterator and array materialization when collecting current taskbar windows.", "P1", "Done", "TaskbarWindowCatalog now builds the current taskbar list directly with a small pre-sized list instead of yielding and materializing an array."],
    [86, "Infrastructure", "Fade duration projection allocation reduction", "Avoid LINQ dictionary projections and duration scans when preparing taskbar fade animations.", "P1", "Done", "TaskbarAppearanceService now builds animation start and duration dictionaries with pre-sized loops and checks for animated durations with a direct scan."],
    [87, "Infrastructure", "Monitor sequence comparison fast path", "Avoid enumerator churn when comparing list-backed monitor profiles during refresh.", "P1", "Done", "MonitorProfile.SequenceMatches now compares IList-backed monitor collections by count and index while preserving the deferred enumerable fallback."],
    [88, "Infrastructure", "Monitor refresh merge loop optimization", "Avoid LINQ projection and delegate searches when merging detected monitor data with saved overrides.", "P1", "Done", "AppState.RefreshMonitors now uses a pre-sized MonitorProfile merge helper with direct device-name lookup loops while preserving saved per-monitor overrides."],
    [89, "Infrastructure", "Monitor catalog ordering allocation reduction", "Avoid LINQ ordering and array materialization when building detected monitor profiles.", "P1", "Done", "MonitorCatalog now builds a pre-sized profile list directly, keeps primary displays first during insertion, and preserves discovery-based display names."],
    [90, "Infrastructure", "Monitor override lookup fast path", "Avoid delegate-based FirstOrDefault searches when applying per-monitor overrides.", "P1", "Done", "AppState.SetMonitorOverride now reuses a list-aware MonitorProfile device lookup helper for saved and live monitor updates while preserving case-insensitive matching."],
    [91, "Infrastructure", "Monitor page refresh scan fast paths", "Avoid delegate-based monitor count and selected-display scans during Monitors page refresh.", "P1", "Done", "Monitors page now uses list-aware MonitorProfile helpers for synced display counts and secondary-or-primary selection."],
    [92, "Infrastructure", "Runtime event list projection allocation reduction", "Avoid LINQ projection allocations when rebuilding version-gated runtime event lists.", "P1", "Done", "Dashboard, Diagnostics, and Monitors now rebuild runtime event list rows with pre-sized loops while preserving existing row text."],
    [93, "Infrastructure", "Taskbar apply handle set pre-sizing", "Avoid avoidable HashSet growth while deduplicating and pruning taskbar handles during apply.", "P1", "Done", "TaskbarAppearanceService now pre-sizes handle sets from known target counts in DistinctByHandle and live-cache pruning."],
    [94, "Infrastructure", "Runtime sensor native check gating", "Skip unneeded native foreground, fullscreen, and taskbar proximity checks based on automation settings.", "P1", "Done", "RuntimeStateSensorService now returns Desktop immediately when automation is disabled and only runs fullscreen or hover/taskbar checks when those policies are enabled."],
    [95, "Infrastructure", "Runtime sensor hover priority short-circuit", "Skip foreground, maximize, and fullscreen native checks when hover proximity already resolves the active runtime trigger.", "P1", "Done", "RuntimeStateSensorService now checks hover proximity first after automation gating and returns Hover immediately when active."],
    [96, "Infrastructure", "Runtime sensor shell class buffer reuse", "Avoid allocating a new class-name StringBuilder during each foreground shell-window check.", "P1", "Done", "RuntimeStateSensorService now reuses one class-name buffer per sensor instance and keeps shell class matching covered by tests."],
]

TAB_COLORS = {
    "Overview": "2878FF",
    "Screens": "23C58E",
    "Features": "FFB454",
    "Backend": "7C3AED",
    "Infrastructure": "64748B",
    "How to use": "101318",
}


def status_counts(rows):
    return {status: sum(1 for row in rows if row[5] == status) for status in STATUSES}


def add_banner(ws, title, color):
    ws.merge_cells("A1:G2")
    cell = ws["A1"]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(color="FFFFFF", bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_table(ws, rows):
    add_banner(ws, ws.title, TAB_COLORS.get(ws.title, "2878FF"))
    for col, header in enumerate(COLUMNS, 1):
        cell = ws.cell(4, col, header)
        cell.fill = PatternFill("solid", fgColor="E8EEF8")
        cell.font = Font(bold=True, color="101318")
        cell.alignment = Alignment(horizontal="center")
    widths = [6, 18, 26, 52, 12, 16, 42]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    thin = Side(style="thin", color="D7DEE8")
    for row_index, row in enumerate(rows, 5):
        for col, value in enumerate(row, 1):
            cell = ws.cell(row_index, col, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
    for extra in range(len(rows) + 5, len(rows) + 15):
        for col in range(1, 8):
            ws.cell(extra, col).fill = PatternFill("solid", fgColor="FFFFFF" if extra % 2 else "F8FAFC")
    ws.freeze_panes = "A5"
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection[0].sqref = "A1"
    ws.sheet_view.selection[0].activeCell = "A1"

    status_validation = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"')
    priority_validation = DataValidation(type="list", formula1='"' + ",".join(PRIORITIES) + '"')
    ws.add_data_validation(status_validation)
    ws.add_data_validation(priority_validation)
    status_validation.add(f"F5:F{len(rows)+14}")
    priority_validation.add(f"E5:E{len(rows)+14}")

    status_colors = {"Done": "DCFCE7", "In Progress": "DBEAFE", "Not Started": "F1F5F9", "Blocked": "FEE2E2"}
    for status, color in status_colors.items():
        ws.conditional_formatting.add(
            f"F5:F{len(rows)+14}",
            FormulaRule(formula=[f'F5="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )
    for priority, color in {"P0": "FEE2E2", "P1": "FEF3C7", "P2": "DBEAFE", "P3": "F1F5F9"}.items():
        ws.conditional_formatting.add(
            f"E5:E{len(rows)+14}",
            FormulaRule(formula=[f'E5="{priority}"'], fill=PatternFill("solid", fgColor=color)),
        )


def build_overview(wb, rows):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = TAB_COLORS["Overview"]
    ws.sheet_view.topLeftCell = "D1"
    ws.sheet_view.selection[0].sqref = "H4"
    ws.sheet_view.selection[0].activeCell = "H4"
    for col in range(1, 16):
        ws.column_dimensions[chr(64 + col)].width = 13
    ws.merge_cells("F2:J3")
    ws["F2"] = "Oxygen Taskbar Project Tracker"
    ws["F2"].font = Font(size=20, bold=True, color="101318")
    ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F4"] = f"LAST_UPDATED: {LAST_UPDATED}"
    ws["F4"].font = Font(color="64748B", italic=True)

    counts = status_counts(rows)
    total = len(rows)
    complete = counts["Done"] / total if total else 0
    cards = [
        ("Total", total),
        ("Done", counts["Done"]),
        ("In Progress", counts["In Progress"]),
        ("Not Started", counts["Not Started"]),
        ("Complete", complete),
    ]
    for i, (label, value) in enumerate(cards):
        col = 4 + i * 2
        ws.merge_cells(start_row=6, start_column=col, end_row=8, end_column=col + 1)
        cell = ws.cell(6, col)
        cell.value = f"{label}\n{value:.0%}" if label == "Complete" else f"{label}\n{value}"
        cell.fill = PatternFill("solid", fgColor="EFF6FF")
        cell.font = Font(size=14, bold=True, color="101318")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["F11"] = "Progress by Area"
    ws["F11"].font = Font(size=14, bold=True)
    areas = sorted(set(row[1] for row in rows))
    for index, area in enumerate(areas, 12):
        area_rows = [row for row in rows if row[1] == area]
        done = sum(1 for row in area_rows if row[5] == "Done")
        ws.cell(index, 6, area)
        ws.cell(index, 7, done / len(area_rows))
        ws.cell(index, 7).number_format = "0%"
    ws.conditional_formatting.add(f"G12:G{11+len(areas)}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="2878FF"))


def build_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = TAB_COLORS[title]
    style_table(ws, rows)


def build_how_to_use(wb):
    ws = wb.create_sheet("How to use")
    ws.sheet_properties.tabColor = TAB_COLORS["How to use"]
    add_banner(ws, "How to use", TAB_COLORS["How to use"])
    ws["A4"] = "Statuses"
    ws["A4"].font = Font(bold=True)
    for index, status in enumerate(STATUSES, 5):
        ws.cell(index, 1, status)
    ws["C4"] = "Priorities"
    ws["C4"].font = Font(bold=True)
    for index, priority in enumerate(PRIORITIES, 5):
        ws.cell(index, 3, priority)
    ws["A11"] = "Update tracker/build_tracker.py, then run python tracker/build_tracker.py. Do not hand-edit generated workbook data."
    ws["A11"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["C"].width = 18


def main():
    root = Path(__file__).resolve().parent
    out = root / "Project-Tracker.xlsx"
    wb = Workbook()
    build_overview(wb, ITEMS)
    for title in ["Screens", "Features", "Backend", "Infrastructure"]:
        build_sheet(wb, title, [row for row in ITEMS if row[1] == title])
    build_how_to_use(wb)
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
